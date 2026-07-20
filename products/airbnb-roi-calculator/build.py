import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
BLUE = Font(name=FONT_NAME, color="0000FF")
BLUE_BOLD = Font(name=FONT_NAME, color="0000FF", bold=True)
BLACK = Font(name=FONT_NAME, color="000000")
BLACK_BOLD = Font(name=FONT_NAME, bold=True)
GREEN = Font(name=FONT_NAME, color="008000")
WHITE_BOLD = Font(name=FONT_NAME, color="FFFFFF", bold=True, size=12)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=16)
SECTION_FONT = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
SECTION_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
LIGHT_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CURRENCY = '$#,##0;($#,##0);"-"'
CURRENCY2 = '$#,##0.00;($#,##0.00);"-"'
PCT = '0.0%;(0.0%);"-"'

wb = openpyxl.Workbook()

# ---------- Instructions sheet ----------
ws = wb.active
ws.title = "Read Me"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 100
ws["A1"] = "Short-Term Rental / Airbnb ROI & Cash Flow Calculator"
ws["A1"].font = TITLE_FONT
ws["A2"] = "How to use this workbook"
ws["A2"].font = BLACK_BOLD
lines = [
    "",
    "1. Go to the 'Inputs' tab. Every cell shaded yellow with blue text is meant for you to edit.",
    "   Replace the example numbers with your own property's numbers.",
    "2. The 'ROI Dashboard' tab updates automatically and shows your key investment metrics:",
    "   Cap Rate, Cash-on-Cash Return, NOI, Monthly & Annual Cash Flow, Break-Even Occupancy,",
    "   and a 5-Year projection.",
    "3. The 'Monthly Cash Flow' tab breaks the first year down month by month, with a seasonality",
    "   slider per month so you can model peak vs. off-peak booking rates.",
    "4. Nothing on the Dashboard or Monthly tabs needs editing -- they are all formulas driven by",
    "   your Inputs. If a number looks wrong, check the matching Inputs cell first.",
    "5. All formulas are live Excel/Google Sheets formulas, not pasted values -- change any input",
    "   and every downstream number recalculates.",
    "",
    "Definitions",
    "  NOI (Net Operating Income) = Gross Rental Income - Operating Expenses (excludes mortgage P&I)",
    "  Cap Rate = NOI / Purchase Price",
    "  Cash-on-Cash Return = Annual Pre-Tax Cash Flow / Total Cash Invested",
    "  Break-Even Occupancy = Rate of nights booked per year needed to cover all cash expenses",
    "",
    "This workbook is a planning tool. It does not constitute financial, tax, or legal advice.",
    "Verify local short-term-rental regulations, licensing, and taxes before purchasing a property.",
]
for i, line in enumerate(lines, start=3):
    c = ws.cell(row=i, column=1, value=line)
    c.font = BLACK_BOLD if line and line[0].isdigit() and line[1] == "." else BLACK
    if line == "Definitions":
        c.font = BLACK_BOLD

for i, line in enumerate(lines, start=3):
    c = ws.cell(row=i, column=1, value=line)
    c.font = BLACK_BOLD if line and line[0].isdigit() and line[1] == "." else BLACK
    if line == "Definitions":
        c.font = BLACK_BOLD

# ---------- Inputs sheet ----------
inp = wb.create_sheet("Inputs")
inp.sheet_view.showGridLines = False
inp.column_dimensions["A"].width = 34
inp.column_dimensions["B"].width = 16
inp.column_dimensions["C"].width = 46

inp["A1"] = "Inputs"
inp["A1"].font = TITLE_FONT
inp["C1"] = "Yellow + blue cells = edit these. Everything else is a formula."
inp["C1"].font = Font(name=FONT_NAME, italic=True, color="666666")

def section(ws_, row, title):
    ws_.cell(row=row, column=1, value=title)
    ws_.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    cell = ws_.cell(row=row, column=1)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(indent=1)
    for col in (2, 3):
        ws_.cell(row=row, column=col).fill = SECTION_FILL

def input_row(ws_, row, label, value, cell_col="B", numfmt=None, note=None):
    ws_.cell(row=row, column=1, value=label).font = BLACK
    c = ws_.cell(row=row, column=2, value=value)
    c.font = BLUE_BOLD
    c.fill = YELLOW_FILL
    c.border = BORDER
    if numfmt:
        c.number_format = numfmt
    if note:
        ws_.cell(row=row, column=3, value=note).font = Font(name=FONT_NAME, italic=True, color="666666", size=9)

def calc_row(ws_, row, label, formula, numfmt=None, bold=False):
    ws_.cell(row=row, column=1, value=label).font = BLACK_BOLD if bold else BLACK
    c = ws_.cell(row=row, column=2, value=formula)
    c.font = BLACK_BOLD if bold else BLACK
    c.border = BORDER
    if numfmt:
        c.number_format = numfmt

r = 3
section(inp, r, "Property & Purchase"); r += 1
input_row(inp, r, "Purchase price", 325000, numfmt=CURRENCY, note="Example: mid-size STR in a secondary market"); PURCHASE_PRICE = r; r += 1
input_row(inp, r, "Closing costs (%  of price)", 0.03, numfmt=PCT); CLOSING_PCT = r; r += 1
input_row(inp, r, "Furnishing & setup budget", 16000, numfmt=CURRENCY, note="Furniture, photography, smart locks, linens, etc."); FURNISH = r; r += 1
input_row(inp, r, "Down payment (%)", 0.20, numfmt=PCT); DOWN_PCT = r; r += 1
input_row(inp, r, "Mortgage interest rate (annual)", 0.065, numfmt=PCT); RATE = r; r += 1
input_row(inp, r, "Loan term (years)", 30, numfmt="0"); TERM = r; r += 1
r += 1

section(inp, r, "Revenue Assumptions"); r += 1
input_row(inp, r, "Average nightly rate", 215, numfmt=CURRENCY); ADR = r; r += 1
input_row(inp, r, "Expected occupancy rate", 0.60, numfmt=PCT, note="Nights booked / nights available, annual average"); OCC = r; r += 1
input_row(inp, r, "Cleaning fee per stay (pass-through)", 90, numfmt=CURRENCY); CLEAN_FEE = r; r += 1
input_row(inp, r, "Average nights per booking", 3.2, numfmt="0.0"); NIGHTS_PER_STAY = r; r += 1
input_row(inp, r, "Other income per year (parking, pet fee, etc.)", 600, numfmt=CURRENCY); OTHER_INC = r; r += 1
input_row(inp, r, "Annual revenue growth rate (for 5-yr projection)", 0.03, numfmt=PCT); REV_GROWTH = r; r += 1
input_row(inp, r, "Annual expense growth rate (for 5-yr projection)", 0.025, numfmt=PCT); EXP_GROWTH = r; r += 1
r += 1

section(inp, r, "Operating Expenses (annual)"); r += 1
input_row(inp, r, "Property management fee (% of revenue)", 0.18, numfmt=PCT); MGMT_PCT = r; r += 1
input_row(inp, r, "Platform fees, e.g. Airbnb/VRBO (% of revenue)", 0.03, numfmt=PCT); PLATFORM_PCT = r; r += 1
input_row(inp, r, "Cleaning & turnover cost (paid to cleaner)", 75, numfmt=CURRENCY, note="Per turnover; usually close to the cleaning fee charged"); CLEAN_COST = r; r += 1
input_row(inp, r, "Property taxes (annual)", 3800, numfmt=CURRENCY); TAXES = r; r += 1
input_row(inp, r, "Insurance, STR policy (annual)", 2000, numfmt=CURRENCY); INSURANCE = r; r += 1
input_row(inp, r, "Utilities (annual)", 3200, numfmt=CURRENCY); UTILITIES = r; r += 1
input_row(inp, r, "Internet, streaming, software (annual)", 900, numfmt=CURRENCY); INTERNET = r; r += 1
input_row(inp, r, "Supplies & consumables (annual)", 1200, numfmt=CURRENCY); SUPPLIES = r; r += 1
input_row(inp, r, "HOA / licensing / permit fees (annual)", 800, numfmt=CURRENCY); HOA = r; r += 1
input_row(inp, r, "Repairs & maintenance reserve (% of revenue)", 0.05, numfmt=PCT); REPAIR_PCT = r; r += 1
r += 1

section(inp, r, "Derived Purchase Numbers"); r += 1
calc_row(inp, r, "Total cash to close (down payment + closing + furnishing)",
         f"=B{PURCHASE_PRICE}*B{DOWN_PCT}+B{PURCHASE_PRICE}*B{CLOSING_PCT}+B{FURNISH}", CURRENCY, bold=True); CASH_TO_CLOSE = r; r += 1
calc_row(inp, r, "Loan amount", f"=B{PURCHASE_PRICE}*(1-B{DOWN_PCT})", CURRENCY); LOAN_AMT = r; r += 1
calc_row(inp, r, "Monthly mortgage payment (P&I)",
         f"=-PMT(B{RATE}/12,B{TERM}*12,B{LOAN_AMT})", CURRENCY, bold=True); MONTHLY_PI = r; r += 1

inp.freeze_panes = "A3"
for row in inp.iter_rows(min_row=3, max_row=r, min_col=1, max_col=3):
    for cell in row:
        if cell.font and cell.font.color and cell.font.color.rgb == "FFFFFFFF":
            continue

IN = "Inputs"

# ---------- ROI Dashboard sheet ----------
db = wb.create_sheet("ROI Dashboard")
db.sheet_view.showGridLines = False
db.column_dimensions["A"].width = 42
db.column_dimensions["B"].width = 16
db.column_dimensions["C"].width = 16
db.column_dimensions["D"].width = 16
db.column_dimensions["E"].width = 16
db.column_dimensions["F"].width = 16
db.column_dimensions["G"].width = 16

db["A1"] = "ROI Dashboard"
db["A1"].font = TITLE_FONT
db["A2"] = "All values below are formulas driven by the Inputs tab. Nothing here needs editing."
db["A2"].font = Font(name=FONT_NAME, italic=True, color="666666")

r = 4
section(db, r, "Revenue"); r += 1
calc_row(db, r, "Annual nights booked", f"=365*{IN}!$B${OCC}", "0"); NIGHTS_BOOKED = r; r += 1
calc_row(db, r, "Number of bookings per year", f"=B{NIGHTS_BOOKED}/{IN}!$B${NIGHTS_PER_STAY}", "0.0"); BOOKINGS = r; r += 1
calc_row(db, r, "Gross rental revenue", f"={IN}!$B${ADR}*B{NIGHTS_BOOKED}", CURRENCY); GROSS_RENT = r; r += 1
calc_row(db, r, "Cleaning fee revenue (pass-through)", f"={IN}!$B${CLEAN_FEE}*B{BOOKINGS}", CURRENCY); CLEAN_REV = r; r += 1
calc_row(db, r, "Other income", f"={IN}!$B${OTHER_INC}", CURRENCY); OTHER_REV = r; r += 1
calc_row(db, r, "Total revenue", f"=SUM(B{GROSS_RENT}:B{OTHER_REV})", CURRENCY, bold=True); TOTAL_REV = r; r += 1
r += 1

section(db, r, "Operating Expenses"); r += 1
calc_row(db, r, "Property management fee", f"={IN}!$B${MGMT_PCT}*B{TOTAL_REV}", CURRENCY); MGMT_FEE = r; r += 1
calc_row(db, r, "Platform fees", f"={IN}!$B${PLATFORM_PCT}*B{TOTAL_REV}", CURRENCY); PLATFORM_FEE = r; r += 1
calc_row(db, r, "Cleaning cost (paid to cleaner)", f"={IN}!$B${CLEAN_COST}*B{BOOKINGS}", CURRENCY); CLEAN_COST_TOT = r; r += 1
calc_row(db, r, "Repairs & maintenance reserve", f"={IN}!$B${REPAIR_PCT}*B{TOTAL_REV}", CURRENCY); REPAIR_TOT = r; r += 1
calc_row(db, r, "Property taxes", f"={IN}!$B${TAXES}", CURRENCY); TAXES_D = r; r += 1
calc_row(db, r, "Insurance", f"={IN}!$B${INSURANCE}", CURRENCY); INS_D = r; r += 1
calc_row(db, r, "Utilities", f"={IN}!$B${UTILITIES}", CURRENCY); UTIL_D = r; r += 1
calc_row(db, r, "Internet, streaming, software", f"={IN}!$B${INTERNET}", CURRENCY); NET_D = r; r += 1
calc_row(db, r, "Supplies & consumables", f"={IN}!$B${SUPPLIES}", CURRENCY); SUP_D = r; r += 1
calc_row(db, r, "HOA / licensing / permits", f"={IN}!$B${HOA}", CURRENCY); HOA_D = r; r += 1
calc_row(db, r, "Total operating expenses", f"=SUM(B{MGMT_FEE}:B{HOA_D})", CURRENCY, bold=True); TOTAL_OPEX = r; r += 1
r += 1

section(db, r, "Net Operating Income & Cash Flow"); r += 1
calc_row(db, r, "Net Operating Income (NOI)", f"=B{TOTAL_REV}-B{TOTAL_OPEX}", CURRENCY, bold=True); NOI = r; r += 1
calc_row(db, r, "Annual debt service (mortgage P&I)", f"={IN}!$B${MONTHLY_PI}*12", CURRENCY); DEBT_SVC = r; r += 1
calc_row(db, r, "Annual cash flow (pre-tax)", f"=B{NOI}-B{DEBT_SVC}", CURRENCY, bold=True); CASH_FLOW = r; r += 1
calc_row(db, r, "Monthly cash flow (pre-tax)", f"=B{CASH_FLOW}/12", CURRENCY, bold=True); MONTHLY_CF = r; r += 1
r += 1

section(db, r, "Investment Returns"); r += 1
calc_row(db, r, "Total cash invested (down payment + closing + furnishing)", f"={IN}!$B${CASH_TO_CLOSE}", CURRENCY); CASH_INV = r; r += 1
calc_row(db, r, "Cap rate  (NOI / purchase price)", f"=B{NOI}/{IN}!$B${PURCHASE_PRICE}", PCT, bold=True); CAP_RATE = r; r += 1
calc_row(db, r, "Cash-on-cash return  (cash flow / cash invested)", f"=B{CASH_FLOW}/B{CASH_INV}", PCT, bold=True); COC = r; r += 1
r += 1

section(db, r, "Break-Even Occupancy (helper calculation)"); r += 1
calc_row(db, r, "Revenue per occupancy point  (365 x (ADR + cleaning fee / nights per stay))",
         f"=365*({IN}!$B${ADR}+{IN}!$B${CLEAN_FEE}/{IN}!$B${NIGHTS_PER_STAY})", CURRENCY2); K1 = r; r += 1
calc_row(db, r, "Margin retained after %-of-revenue costs",
         f"=1-({IN}!$B${MGMT_PCT}+{IN}!$B${PLATFORM_PCT}+{IN}!$B${REPAIR_PCT})", PCT); MARGIN = r; r += 1
calc_row(db, r, "Variable cleaning cost per occupancy point  (365 x cleaning cost / nights per stay)",
         f"=365*{IN}!$B${CLEAN_COST}/{IN}!$B${NIGHTS_PER_STAY}", CURRENCY2); VARCLEAN = r; r += 1
calc_row(db, r, "Fixed annual costs  (taxes+insurance+utilities+internet+supplies+HOA+debt service)",
         f"=B{TAXES_D}+B{INS_D}+B{UTIL_D}+B{NET_D}+B{SUP_D}+B{HOA_D}+B{DEBT_SVC}", CURRENCY); FIXED_TOT = r; r += 1
calc_row(db, r, "Break-even occupancy rate",
         f"=(B{FIXED_TOT}-B{MARGIN}*{IN}!$B${OTHER_INC})/(B{MARGIN}*B{K1}-B{VARCLEAN})", PCT, bold=True); BREAKEVEN = r; r += 1
calc_row(db, r, "Break-even nights booked per year", f"=B{BREAKEVEN}*365", "0", bold=True); BREAKEVEN_NIGHTS = r; r += 1
r += 1

section(db, r, "5-Year Projection"); r += 1
db.cell(row=r, column=1, value="").font = BLACK
for i, yr in enumerate(range(1, 6)):
    c = db.cell(row=r, column=2 + i, value=f"Year {yr}")
    c.font = BLACK_BOLD
    c.fill = LIGHT_FILL
    c.alignment = Alignment(horizontal="center")
    c.border = BORDER
YEAR_HDR = r; r += 1

def year_formula(base_cell_ref, growth_input_row, col_idx):
    if col_idx == 0:
        return f"={base_cell_ref}"
    prev_col = get_column_letter(1 + col_idx)
    return f"={prev_col}{{row}}*(1+{IN}!$B${growth_input_row})"

REV_ROW = r
db.cell(row=r, column=1, value="Total revenue").font = BLACK_BOLD
for i in range(5):
    col = get_column_letter(2 + i)
    if i == 0:
        f = f"=B{TOTAL_REV}"
    else:
        prev = get_column_letter(1 + i)
        f = f"={prev}{r}*(1+{IN}!$B${REV_GROWTH})"
    c = db.cell(row=r, column=2 + i, value=f)
    c.number_format = CURRENCY
    c.font = BLACK_BOLD
    c.border = BORDER
r += 1

EXP_ROW = r
db.cell(row=r, column=1, value="Total operating expenses").font = BLACK
for i in range(5):
    if i == 0:
        f = f"=B{TOTAL_OPEX}"
    else:
        prev = get_column_letter(1 + i)
        f = f"={prev}{r}*(1+{IN}!$B${EXP_GROWTH})"
    c = db.cell(row=r, column=2 + i, value=f)
    c.number_format = CURRENCY
    c.font = BLACK
    c.border = BORDER
r += 1

NOI_ROW = r
db.cell(row=r, column=1, value="NOI").font = BLACK_BOLD
for i in range(5):
    col = get_column_letter(2 + i)
    c = db.cell(row=r, column=2 + i, value=f"={col}{REV_ROW}-{col}{EXP_ROW}")
    c.number_format = CURRENCY
    c.font = BLACK_BOLD
    c.border = BORDER
r += 1

DEBT_ROW = r
db.cell(row=r, column=1, value="Annual debt service").font = BLACK
for i in range(5):
    c = db.cell(row=r, column=2 + i, value=f"=$B${DEBT_SVC}")
    c.number_format = CURRENCY
    c.font = BLACK
    c.border = BORDER
r += 1

CF_ROW = r
db.cell(row=r, column=1, value="Cash flow (pre-tax)").font = BLACK_BOLD
for i in range(5):
    col = get_column_letter(2 + i)
    c = db.cell(row=r, column=2 + i, value=f"={col}{NOI_ROW}-{col}{DEBT_ROW}")
    c.number_format = CURRENCY
    c.font = BLACK_BOLD
    c.border = BORDER
r += 1

CUM_ROW = r
db.cell(row=r, column=1, value="Cumulative cash flow").font = BLACK_BOLD
for i in range(5):
    col = get_column_letter(2 + i)
    if i == 0:
        f = f"={col}{CF_ROW}"
    else:
        prev = get_column_letter(1 + i)
        f = f"={prev}{CUM_ROW}+{col}{CF_ROW}"
    c = db.cell(row=r, column=2 + i, value=f)
    c.number_format = CURRENCY
    c.font = BLACK_BOLD
    c.fill = LIGHT_FILL
    c.border = BORDER
r += 1

db.freeze_panes = "A4"

# ---------- Monthly Cash Flow sheet ----------
mo = wb.create_sheet("Monthly Cash Flow")
mo.sheet_view.showGridLines = False
mo.column_dimensions["A"].width = 30
for col in "BCDEFGHIJKLMN":
    mo.column_dimensions[col].width = 11

mo["A1"] = "Monthly Cash Flow (Year 1)"
mo["A1"].font = TITLE_FONT
mo["A2"] = "Edit the seasonality factor per month (yellow). 1.00 = average month; 1.30 = 30% above average."
mo["A2"].font = Font(name=FONT_NAME, italic=True, color="666666")

months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
seasonality_defaults = [0.75, 0.80, 0.95, 1.05, 1.10, 1.25, 1.35, 1.30, 1.05, 0.95, 0.80, 0.85]

r0 = 4
mo.cell(row=r0, column=1, value="Month").font = BLACK_BOLD
for i, m in enumerate(months):
    c = mo.cell(row=r0, column=2 + i, value=m)
    c.font = BLACK_BOLD
    c.fill = LIGHT_FILL
    c.alignment = Alignment(horizontal="center")
    c.border = BORDER
c = mo.cell(row=r0, column=14, value="Total")
c.font = BLACK_BOLD; c.fill = LIGHT_FILL; c.border = BORDER
SEASON_ROW = r0 + 1
mo.cell(row=SEASON_ROW, column=1, value="Seasonality factor").font = BLACK
for i, v in enumerate(seasonality_defaults):
    c = mo.cell(row=SEASON_ROW, column=2 + i, value=v)
    c.font = BLUE_BOLD
    c.fill = YELLOW_FILL
    c.number_format = "0.00"
    c.border = BORDER
c = mo.cell(row=SEASON_ROW, column=14, value=f"=AVERAGE(B{SEASON_ROW}:M{SEASON_ROW})")
c.font = BLACK; c.number_format = "0.00"; c.border = BORDER
NORMALIZER_ROW = SEASON_ROW  # reuse average as normalizer denominator inline in formulas below

REV_M_ROW = SEASON_ROW + 1
mo.cell(row=REV_M_ROW, column=1, value="Revenue").font = BLACK_BOLD
for i in range(12):
    col = get_column_letter(2 + i)
    f = f"={col}{SEASON_ROW}/$N${SEASON_ROW}*'ROI Dashboard'!$B${TOTAL_REV}/12"
    c = mo.cell(row=REV_M_ROW, column=2 + i, value=f)
    c.number_format = CURRENCY
    c.font = BLACK
    c.border = BORDER
c = mo.cell(row=REV_M_ROW, column=14, value=f"=SUM(B{REV_M_ROW}:M{REV_M_ROW})")
c.font = BLACK_BOLD; c.number_format = CURRENCY; c.border = BORDER

EXP_M_ROW = REV_M_ROW + 1
mo.cell(row=EXP_M_ROW, column=1, value="Operating expenses").font = BLACK
for i in range(12):
    col = get_column_letter(2 + i)
    f = f"={col}{SEASON_ROW}/$N${SEASON_ROW}*'ROI Dashboard'!$B${TOTAL_OPEX}/12"
    c = mo.cell(row=EXP_M_ROW, column=2 + i, value=f)
    c.number_format = CURRENCY
    c.font = BLACK
    c.border = BORDER
c = mo.cell(row=EXP_M_ROW, column=14, value=f"=SUM(B{EXP_M_ROW}:M{EXP_M_ROW})")
c.font = BLACK_BOLD; c.number_format = CURRENCY; c.border = BORDER

DEBT_M_ROW = EXP_M_ROW + 1
mo.cell(row=DEBT_M_ROW, column=1, value="Mortgage payment (P&I)").font = BLACK
for i in range(12):
    c = mo.cell(row=DEBT_M_ROW, column=2 + i, value=f"={IN}!$B${MONTHLY_PI}")
    c.number_format = CURRENCY
    c.font = BLACK
    c.border = BORDER
c = mo.cell(row=DEBT_M_ROW, column=14, value=f"=SUM(B{DEBT_M_ROW}:M{DEBT_M_ROW})")
c.font = BLACK_BOLD; c.number_format = CURRENCY; c.border = BORDER

CF_M_ROW = DEBT_M_ROW + 1
mo.cell(row=CF_M_ROW, column=1, value="Net cash flow").font = BLACK_BOLD
for i in range(12):
    col = get_column_letter(2 + i)
    c = mo.cell(row=CF_M_ROW, column=2 + i, value=f"={col}{REV_M_ROW}-{col}{EXP_M_ROW}-{col}{DEBT_M_ROW}")
    c.number_format = CURRENCY
    c.font = BLACK_BOLD
    c.border = BORDER
c = mo.cell(row=CF_M_ROW, column=14, value=f"=SUM(B{CF_M_ROW}:M{CF_M_ROW})")
c.font = BLACK_BOLD; c.number_format = CURRENCY; c.border = BORDER

CUM_M_ROW = CF_M_ROW + 1
mo.cell(row=CUM_M_ROW, column=1, value="Cumulative cash flow").font = BLACK_BOLD
for i in range(12):
    col = get_column_letter(2 + i)
    if i == 0:
        f = f"={col}{CF_M_ROW}"
    else:
        prev = get_column_letter(1 + i)
        f = f"={prev}{CUM_M_ROW}+{col}{CF_M_ROW}"
    c = mo.cell(row=CUM_M_ROW, column=2 + i, value=f)
    c.number_format = CURRENCY
    c.font = BLACK_BOLD
    c.fill = LIGHT_FILL
    c.border = BORDER

mo.freeze_panes = "B5"

wb.active = 1  # open on Inputs tab by default
wb.save("Airbnb_STR_ROI_Calculator.xlsx")
print("monthly sheet done")
