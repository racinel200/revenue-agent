import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
BLUE_BOLD = Font(name=FONT_NAME, color="0000FF", bold=True)
BLACK = Font(name=FONT_NAME, color="000000")
BLACK_BOLD = Font(name=FONT_NAME, bold=True)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=16)
SECTION_FONT = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
SECTION_FILL = PatternFill(start_color="8E4B6B", end_color="8E4B6B", fill_type="solid")
LIGHT_FILL = PatternFill(start_color="F3D9E2", end_color="F3D9E2", fill_type="solid")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CURRENCY = '$#,##0;($#,##0);"-"'
PCT = '0.0%;(0.0%);"-"'

CATEGORIES = [
    "Venue & Rentals", "Catering & Bar", "Photography & Video", "Wedding Attire",
    "Hair & Makeup", "Flowers & Decor", "Music & Entertainment",
    "Invitations & Stationery", "Wedding Cake & Desserts", "Transportation",
    "Rings", "Officiant", "Favors & Gifts", "Miscellaneous / Buffer",
]
PLANNED = [9000, 7500, 3000, 1500, 600, 2100, 1800, 450, 600, 450, 1500, 400, 450, 650]

# vendor rows: category, vendor name, contact, contract total, deposit amount,
# deposit due, deposit paid, deposit due date offset text, balance due date, balance paid, notes
VENDORS = [
    ("Venue & Rentals", "Sunset Barn Venue", "booking@sunsetbarn.example", 9400, 2800, "Booked", "3 mo before"),
    ("Catering & Bar", "Harvest Table Catering", "events@harvesttable.example", 7200, 2200, "Booked", "1 mo before"),
    ("Photography & Video", "Golden Hour Photo Co.", "hello@goldenhourphoto.example", 3200, 1000, "Booked", "Day-of"),
    ("Wedding Attire", "Bridal Boutique + Men's Wearhouse", "-", 1450, 1450, "Paid in full", "-"),
    ("Hair & Makeup", "Glow Beauty Studio", "book@glowbeauty.example", 600, 150, "Booked", "Day-of"),
    ("Flowers & Decor", "Bloom & Co Florals", "orders@bloomandco.example", 2250, 675, "Booked", "2 wk before"),
    ("Music & Entertainment", "DJ Spinlist", "dj@spinlist.example", 1800, 500, "Booked", "Day-of"),
    ("Invitations & Stationery", "Paper & Pine Design", "studio@paperpine.example", 420, 420, "Paid in full", "-"),
    ("Wedding Cake & Desserts", "Sweet Layer Bakery", "orders@sweetlayer.example", 650, 200, "Booked", "1 wk before"),
    ("Transportation", "Classic Car Rentals", "rentals@classiccars.example", 400, 100, "Booked", "Day-of"),
    ("Rings", "Heirloom Jewelers", "sales@heirloomjewelers.example", 1600, 1600, "Paid in full", "-"),
    ("Officiant", "Rev. J. Alvarez", "j.alvarez@example.com", 400, 100, "Booked", "Day-of"),
    ("Favors & Gifts", "Little Things Co.", "orders@littlethings.example", 480, 480, "Paid in full", "-"),
]

wb = openpyxl.Workbook()

# ---------- Read Me ----------
ws = wb.active
ws.title = "Read Me"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 100
ws["A1"] = "Wedding Budget & Vendor Payment Tracker"
ws["A1"].font = TITLE_FONT
ws["A2"] = "How to use this workbook"
ws["A2"].font = BLACK_BOLD
lines = [
    "",
    "1. Go to the 'Budget Overview' tab. Set your total wedding budget in the yellow cell at",
    "   the top, then edit the 'Planned Budget' column (yellow) for each category.",
    "2. Go to the 'Vendor & Payment Tracker' tab. Add one row per vendor: contract total,",
    "   deposit amount, due dates, and paid status. Match the Category column exactly to a",
    "   category name from Budget Overview so the totals roll up correctly.",
    "3. The 'Actual Spent' and 'Difference' columns on Budget Overview update automatically —",
    "   they sum every vendor row tagged with that category. Nothing there needs editing.",
    "4. 'Balance Due' on the Vendor tab is Contract Total minus Deposit Amount — a live formula,",
    "   not a typed number.",
    "5. Replace the example rows with your own vendors; delete the examples once you no longer",
    "   need them as a reference.",
    "",
    "This workbook is a planning tool. It does not constitute financial advice.",
]
for i, line in enumerate(lines, start=3):
    c = ws.cell(row=i, column=1, value=line)
    c.font = BLACK_BOLD if line[:2].strip().isdigit() else BLACK

def section(ws_, row, title, ncols):
    ws_.cell(row=row, column=1, value=title)
    ws_.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws_.cell(row=row, column=1)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(indent=1)
    for col in range(2, ncols + 1):
        ws_.cell(row=row, column=col).fill = SECTION_FILL

# ---------- Budget Overview ----------
bo = wb.create_sheet("Budget Overview")
bo.sheet_view.showGridLines = False
bo.column_dimensions["A"].width = 30
for col in "BCDE":
    bo.column_dimensions[col].width = 16

bo["A1"] = "Budget Overview"
bo["A1"].font = TITLE_FONT

bo["A3"] = "Total wedding budget"
bo["A3"].font = BLACK_BOLD
c = bo.cell(row=3, column=2, value=30000)
c.font = BLUE_BOLD
c.fill = YELLOW_FILL
c.number_format = CURRENCY
c.border = BORDER
TOTAL_BUDGET_CELL = "B3"

header_row = 5
section(bo, header_row, "By Category", 5)
header_row += 1
headers = ["Category", "Planned Budget", "Actual Spent", "Difference (Planned - Actual)", "% of Total Budget"]
for i, h in enumerate(headers):
    c = bo.cell(row=header_row, column=1 + i, value=h)
    c.font = BLACK_BOLD
    c.fill = LIGHT_FILL
    c.border = BORDER
    c.alignment = Alignment(wrap_text=True, vertical="center")

first_data_row = header_row + 1
VT = "'Vendor & Payment Tracker'"
# Vendor tracker data rows are written further down (VENDOR_FIRST_ROW..VENDOR_LAST_ROW);
# computed after we know the sheet layout, referenced here via placeholders filled below.
category_rows = {}
for i, (cat, planned) in enumerate(zip(CATEGORIES, PLANNED)):
    row = first_data_row + i
    category_rows[cat] = row
    bo.cell(row=row, column=1, value=cat).font = BLACK
    bo.cell(row=row, column=1).border = BORDER
    c = bo.cell(row=row, column=2, value=planned)
    c.font = BLUE_BOLD
    c.fill = YELLOW_FILL
    c.number_format = CURRENCY
    c.border = BORDER

last_data_row = first_data_row + len(CATEGORIES) - 1
total_row = last_data_row + 1
bo.cell(row=total_row, column=1, value="Total").font = BLACK_BOLD
c = bo.cell(row=total_row, column=2, value=f"=SUM(B{first_data_row}:B{last_data_row})")
c.font = BLACK_BOLD
c.number_format = CURRENCY
c.border = BORDER
bo.freeze_panes = f"A{first_data_row}"

# ---------- Vendor & Payment Tracker ----------
vt = wb.create_sheet("Vendor & Payment Tracker")
vt.sheet_view.showGridLines = False
vt.column_dimensions["A"].width = 24
vt.column_dimensions["B"].width = 26
vt.column_dimensions["C"].width = 26
for col in "DEFGHIJK":
    vt.column_dimensions[col].width = 14
vt.column_dimensions["K"].width = 22

vt["A1"] = "Vendor & Payment Tracker"
vt["A1"].font = TITLE_FONT

vheader_row = 3
vheaders = ["Category", "Vendor Name", "Contact", "Contract Total", "Deposit Amount",
            "Deposit Status", "Deposit Due", "Balance Due", "Balance Due Date",
            "Balance Paid?", "Notes"]
for i, h in enumerate(vheaders):
    c = vt.cell(row=vheader_row, column=1 + i, value=h)
    c.font = BLACK_BOLD
    c.fill = LIGHT_FILL
    c.border = BORDER
    c.alignment = Alignment(wrap_text=True, vertical="center")

VENDOR_FIRST_ROW = vheader_row + 1
for i, (cat, name, contact, contract, deposit, dep_status, due) in enumerate(VENDORS):
    row = VENDOR_FIRST_ROW + i
    vt.cell(row=row, column=1, value=cat).font = BLACK
    vt.cell(row=row, column=2, value=name).font = BLACK
    vt.cell(row=row, column=3, value=contact).font = BLACK
    c4 = vt.cell(row=row, column=4, value=contract)
    c4.font = BLUE_BOLD; c4.fill = YELLOW_FILL; c4.number_format = CURRENCY
    c5 = vt.cell(row=row, column=5, value=deposit)
    c5.font = BLUE_BOLD; c5.fill = YELLOW_FILL; c5.number_format = CURRENCY
    vt.cell(row=row, column=6, value=dep_status).font = BLACK
    vt.cell(row=row, column=7, value=due).font = BLACK
    c8 = vt.cell(row=row, column=8, value=f"=D{row}-E{row}")
    c8.font = BLACK; c8.number_format = CURRENCY
    vt.cell(row=row, column=9, value="-" if dep_status == "Paid in full" else "See deposit due").font = BLACK
    vt.cell(row=row, column=10, value="Yes" if dep_status == "Paid in full" else "No").font = BLACK
    vt.cell(row=row, column=11, value="").font = BLACK
    for col in range(1, 12):
        vt.cell(row=row, column=col).border = BORDER

VENDOR_LAST_ROW = VENDOR_FIRST_ROW + len(VENDORS) - 1
vtotal_row = VENDOR_LAST_ROW + 1
vt.cell(row=vtotal_row, column=1, value="Total").font = BLACK_BOLD
for col, letter in ((4, "D"), (5, "E"), (8, "H")):
    c = vt.cell(row=vtotal_row, column=col, value=f"=SUM({letter}{VENDOR_FIRST_ROW}:{letter}{VENDOR_LAST_ROW})")
    c.font = BLACK_BOLD
    c.number_format = CURRENCY
    c.border = BORDER
vt.freeze_panes = f"A{VENDOR_FIRST_ROW}"

# ---------- back-fill Budget Overview formulas that depend on the Vendor Tracker range ----------
for cat, row in category_rows.items():
    c = bo.cell(row=row, column=3,
                value=f"=SUMIF({VT}!$A${VENDOR_FIRST_ROW}:$A${VENDOR_LAST_ROW},A{row},"
                      f"{VT}!$D${VENDOR_FIRST_ROW}:$D${VENDOR_LAST_ROW})")
    c.font = BLACK; c.number_format = CURRENCY; c.border = BORDER
    c = bo.cell(row=row, column=4, value=f"=B{row}-C{row}")
    c.font = BLACK; c.number_format = CURRENCY; c.border = BORDER
    c = bo.cell(row=row, column=5, value=f"=B{row}/{TOTAL_BUDGET_CELL}")
    c.font = BLACK; c.number_format = PCT; c.border = BORDER

c = bo.cell(row=total_row, column=3, value=f"=SUM(C{first_data_row}:C{last_data_row})")
c.font = BLACK_BOLD; c.number_format = CURRENCY; c.border = BORDER
c = bo.cell(row=total_row, column=4, value=f"=SUM(D{first_data_row}:D{last_data_row})")
c.font = BLACK_BOLD; c.number_format = CURRENCY; c.border = BORDER
c = bo.cell(row=total_row, column=5, value=f"=SUM(E{first_data_row}:E{last_data_row})")
c.font = BLACK_BOLD; c.number_format = PCT; c.border = BORDER

# remaining-vs-budget summary
summary_row = total_row + 2
bo.cell(row=summary_row, column=1, value="Remaining vs. total budget").font = BLACK_BOLD
c = bo.cell(row=summary_row, column=2, value=f"={TOTAL_BUDGET_CELL}-C{total_row}")
c.font = BLACK_BOLD; c.number_format = CURRENCY; c.border = BORDER

wb.active = 1  # open on Budget Overview tab by default
wb.save("Wedding_Budget_Vendor_Tracker.xlsx")
print("build complete")
print("category_rows:", category_rows)
print("VENDOR_FIRST_ROW", VENDOR_FIRST_ROW, "VENDOR_LAST_ROW", VENDOR_LAST_ROW)
print("first_data_row", first_data_row, "last_data_row", last_data_row, "total_row", total_row)
